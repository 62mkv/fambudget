import logging
from datetime import timedelta

from openpyxl.reader import excel

from cbrapi.client import CbrApiClient, RATE, DATE
from constants import RRU
from dbtables.repository import CurrencyRates


class CurrencyRatesImporter:
    def __init__(self, dbfile, other_currency):
        logging.info('CurrencyRatesImporter is created for %s', other_currency)
        self.rates_table = CurrencyRates(dbfile)
        self.base_currency = RRU
        self.other_currency = other_currency

    # TODO: fix to make it consistent with filling gaps procedure!
    def read_values_from_xls(self, filename):
        """
        In order to import data, download XLS file from this web page:
        https://www.cbr.ru/currency_base/dynamics/?UniDbQuery.Posted=True&UniDbQuery.mode=1&UniDbQuery.date_req1=&UniDbQuery.date_req2=&UniDbQuery.VAL_NM_RQ=R01239

        :param filename: like source-data/RC_F01_01_2008_T05_07_2018.xlsx
        :return:
        """
        raise BaseException('This method has unresolved issues and needs to be fixed!')

        wb = excel.load_workbook(filename=filename)

        ws = wb.worksheets[0]

        row = 2

        names = ('base_currency', 'other_currency', 'rate', 'date')

        while ws.cell(row, 1).value is not None:
            current_date = ws.cell(row, 2).value
            current_rate = ws.cell(row, 3).value
            row += 1
            values = (self.base_currency, self.other_currency, current_rate, current_date)
            record = dict(zip(names, values))
            yield record

    def read_rates_from_cbr(self, till_date):
        logging.info('Updating currency rates for %s till %s', self.other_currency, till_date)
        last_present_date = self.rates_table.get_latest_record_date()

        names = ('base_currency', 'other_currency', 'rate', 'date')
        rates = CbrApiClient().retrieve_rates_for_range(last_present_date + timedelta(days=1), till_date,
                                                        self.other_currency)
        rates.sort(key=lambda x: x[DATE])

        for rate_record in rates:
            values = (self.base_currency, self.other_currency, rate_record[RATE], rate_record[DATE])
            record = dict(zip(names, values))
            yield record

    def fill_missing_dates(self, sequence):
        last_date = self.rates_table.get_latest_record_date()
        last_rate = self.rates_table.get_rate_for_date(self.base_currency, self.other_currency, last_date)
        last_date = last_date + timedelta(days=1)

        for record in sequence:
            new_record = record.copy()
            record_date = record['date']
            while record_date > last_date:
                new_record['date'] = last_date
                new_record['rate'] = last_rate
                last_date += timedelta(days=1)
                yield new_record
            last_rate = record['rate']
            last_date = record_date + timedelta(days=1)
            yield record

    def import_rates_from_xls(self, filename):
        self.rates_table.fill_table_with_records(
            self.fill_missing_dates(
                self.read_values_from_xls(filename)
            )
        )

    def import_rates_from_cbr(self, till_date):
        self.rates_table.fill_table_with_records(
            self.fill_missing_dates(
                self.read_rates_from_cbr(till_date)
            )
        )
