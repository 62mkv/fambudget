# DONE: create a table for exchange rates
# DONE: implement import from xlsx
# DONE: update import fambudget to include also positive amounts
# DONE: create a table for multi-currency transactions (separate row per currency)
# 1) create table with columns: row_index, subject, category, subcount1, subcount2 (row_index is a primary key)
# 2) create table with columns: currency, amount, row_index
# DONE: create table for aggregated data
# 3) create table with columns: row_index, amount_rub, amount_eur
"""
Currently, script was used to populate data:

  insert into spendings (row_index, spent_on, subject, category, subcount1, subcount2)
  select distinct row_index, spent_on, subject, category, subcount1, subcount2 from fambudget

  insert into spending_amounts (row_index, amount, currency)
  select row_index, amount, currency from fambudget
"""
# DONE: update model.json to support normalized tables (spendings/spending_amounts)
# DONE: update model.json: add cube for aggregated (multi-currency) table
# DONE: implement import from cbr API
from datetime import timedelta

from openpyxl.reader import excel

from constants import EUR, RRU
from dbtables import repository


def read_values(filename):
    """
    In order to import data, download XLS file from this web page:
    https://www.cbr.ru/currency_base/dynamics/?UniDbQuery.Posted=True&UniDbQuery.mode=1&UniDbQuery.date_req1=&UniDbQuery.date_req2=&UniDbQuery.VAL_NM_RQ=R01239

    :param filename: like source-data/RC_F01_01_2008_T05_07_2018.xlsx
    :return:
    """
    wb = excel.load_workbook(filename=filename)

    ws = wb.worksheets[0]

    row = 2

    current_date = ws.cell(row, 2).value
    current_rate = ws.cell(row, 3).value

    while ws.cell(row, 1).value is not None:
        next_date = ws.cell(row, 2).value
        while next_date > current_date:
            names = ('base_currency', 'other_currency', 'rate', 'date')
            values = ( RRU, EUR, current_rate, current_date)
            record = dict(zip(names, values))
            yield record
            current_date += timedelta(days=1)

        current_rate = ws.cell(row, 3).value
        row += 1


table = repository.CurrencyRates('sqlite:///data/data.sqlite')
# table.fill_table_with_records(read_values("source-data/RC_F01_01_2008_T05_07_2018.xlsx"))
# table.fill_table_with_records(read_values("source-data/RC_F05_07_2018_T12_01_2019.xlsx"))
# table.fill_table_with_records(read_values("source-data/RC_F11_01_2019_T29_01_2019.xlsx"))
table.fill_table_with_records(read_values("source-data/RC_F29_01_2019_T16_02_2019.xlsx"))
